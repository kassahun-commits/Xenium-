import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE = '../EDIT  Philipp Alcohol proteome Simplified - Jan 2024  copy.xlsx'
OUT  = 'Venn_AcuteWithdrawal_genelists.xlsx'

CORR_THRESH = 3.3
FC_THRESH   = 0.5

SHEETS = [
    ('Membrane',        'Membrane',        'Memb_N-',  'Memb_I-',  'Memb_PA-',  'all'),
    ('Cytosol',         'Cytosol',         'Cyto_N-',  'Cyto_I-',  'Cyto_PA-',  'all'),
    ('Chromatin',       'Chromatin',       'Chrom_N-', 'Chrom_I-', 'Chrom_PA-', 'keep_review'),
    ('Soluble nuclear', 'Soluble nuclear', 'Nuc_N-',   'Nuc_I-',   'Nuc_PA-',   'keep_review'),
]

# AW-M-3 excluded as a global outlier
AW_SUFFIXES = ['AW-F-1', 'AW-F-2', 'AW-M-1', 'AW-M-2']

# ── Venn region labels ──
REGIONS = [
    ('AW only',              lambda A,B,C: A - B - C),
    ('Intox only',           lambda A,B,C: B - A - C),
    ('PA only',              lambda A,B,C: C - A - B),
    ('AW + Intox',           lambda A,B,C: (A & B) - C),
    ('AW + PA',              lambda A,B,C: (A & C) - B),
    ('Intox + PA',           lambda A,B,C: (B & C) - A),
    ('AW + Intox + PA',      lambda A,B,C: A & B & C),
]

def calc_stats(df, naive_cols, cond_cols):
    naive = df[naive_cols].apply(pd.to_numeric, errors='coerce')
    cond  = df[cond_cols].apply(pd.to_numeric, errors='coerce')
    fc    = cond.mean(axis=1) - naive.mean(axis=1)
    p_vals = []
    for i in range(len(df)):
        n = naive.iloc[i].dropna().values
        c = cond.iloc[i].dropna().values
        if len(n) >= 2 and len(c) >= 2:
            _, p = stats.ttest_ind(n, c, equal_var=False)
            p_vals.append(p)
        else:
            p_vals.append(np.nan)
    p_vals    = pd.Series(p_vals, index=df.index)
    valid     = p_vals.notna()
    ranks     = p_vals[valid].rank(ascending=False)
    corrected = pd.Series(np.nan, index=df.index)
    corrected[valid] = -np.log2(np.minimum(1, p_vals[valid] * valid.sum() / ranks))
    return fc, corrected

def get_sig_genes(df, naive_cols, cond_cols, direction):
    fc, corr = calc_stats(df, naive_cols, cond_cols)
    mask = fc.notna() & corr.notna() & (corr > CORR_THRESH)
    if direction == 'up':
        sig = mask & (fc > FC_THRESH)
    else:
        sig = mask & (fc < -FC_THRESH)
    # Return set of genes with their FC and corrected p
    sig_df = df.loc[sig, ['Gene symbol']].copy()
    sig_df['FC']          = fc[sig].values
    sig_df['Corrected_p'] = corr[sig].values
    sig_df['Gene symbol'] = sig_df['Gene symbol'].astype(str)
    return sig_df.set_index('Gene symbol')

# ── Styling helpers ──
HEADER_FILLS = {
    'up':   PatternFill('solid', fgColor='F4C2C2'),   # light pink
    'down': PatternFill('solid', fgColor='C2D4F4'),   # light blue
}
REGION_FILL  = PatternFill('solid', fgColor='F0F0F0')
BOLD         = Font(bold=True, name='Arial', size=11)
NORMAL       = Font(name='Arial', size=10)
CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin         = Side(style='thin', color='BBBBBB')
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_venn_sheet(ws, label, direction, A_df, B_df, C_df,
                     a_label, b_label, c_label):
    """Write one Venn gene-list sheet."""
    A = set(A_df.index)
    B = set(B_df.index)
    C = set(C_df.index)

    # Build columns: for each region, list genes + their FC from whichever set
    # they come from (AW set takes priority for FC values)
    col = 1
    for region_name, region_fn in REGIONS:
        genes = sorted(region_fn(A, B, C))

        # Header row 1: region name
        hdr = ws.cell(row=1, column=col, value=f'{region_name}  (n={len(genes)})')
        hdr.font        = Font(bold=True, name='Arial', size=11)
        hdr.fill        = HEADER_FILLS[direction]
        hdr.alignment   = CENTER
        hdr.border      = BORDER

        # Header row 2: column sub-labels
        for offset, sub in enumerate(['Gene', 'AW FC', 'AW corr-p']):
            c = ws.cell(row=2, column=col+offset, value=sub)
            c.font      = Font(bold=True, name='Arial', size=10)
            c.fill      = REGION_FILL
            c.alignment = CENTER
            c.border    = BORDER

        # Merge region header across 3 columns
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1,   end_column=col+2)

        # Data rows
        for r, gene in enumerate(genes, start=3):
            if gene in A_df.index:
                row_fc   = A_df.loc[gene, 'FC']
                row_corr = A_df.loc[gene, 'Corrected_p']
                # Handle duplicate gene names (take first value)
                fc_val   = round(float(row_fc.iloc[0])   if hasattr(row_fc,   'iloc') else float(row_fc),   3)
                corr_val = round(float(row_corr.iloc[0]) if hasattr(row_corr, 'iloc') else float(row_corr), 3)
            else:
                fc_val, corr_val = '', ''
            for offset, val in enumerate([gene, fc_val, corr_val]):
                c = ws.cell(row=r, column=col+offset, value=val)
                c.font      = NORMAL
                c.alignment = Alignment(horizontal='center')
                c.border    = BORDER

        # Column widths
        ws.column_dimensions[get_column_letter(col)].width   = 14
        ws.column_dimensions[get_column_letter(col+1)].width = 9
        ws.column_dimensions[get_column_letter(col+2)].width = 11

        col += 4   # gap column between regions

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 20

# ── Build workbook ──
wb = Workbook()
wb.remove(wb.active)   # remove default sheet

for label, excel_name, n_pat, i_pat, pa_pat, fmode in SHEETS:
    raw = pd.read_excel(FILE, sheet_name=excel_name, header=0)
    if fmode == 'keep_review':
        df = raw[raw['Filter'].isin(['Keep', 'Review'])].reset_index(drop=True)
    else:
        df = raw.copy().reset_index(drop=True)

    prefix = n_pat.split('_N-')[0] + '_'
    naive_cols = [c for c in df.columns if n_pat  in str(c)]
    i_cols     = [c for c in df.columns if i_pat  in str(c)]
    aw_cols    = [c for c in df.columns if any(prefix + s in str(c) for s in AW_SUFFIXES)]
    pa_cols    = [c for c in df.columns if pa_pat in str(c)]

    # UP sets
    aw_up_df = get_sig_genes(df, naive_cols, aw_cols, 'up')
    i_up_df  = get_sig_genes(df, naive_cols, i_cols,  'up')
    pa_up_df = get_sig_genes(df, naive_cols, pa_cols, 'up')

    # DOWN sets
    aw_dn_df = get_sig_genes(df, naive_cols, aw_cols, 'down')
    i_dn_df  = get_sig_genes(df, naive_cols, i_cols,  'down')
    pa_dn_df = get_sig_genes(df, naive_cols, pa_cols, 'down')

    short = label.replace('Soluble nuclear', 'SolNuc').replace(' ', '')[:12]

    # UP sheet
    ws_up = wb.create_sheet(title=f'{short}_UP')
    ws_up.sheet_properties.tabColor = 'F4C2C2'
    write_venn_sheet(ws_up, label, 'up',
                     aw_up_df, i_up_df, pa_up_df,
                     'AW Increased', 'Intox. Increased', 'PA Increased')

    # DOWN sheet
    ws_dn = wb.create_sheet(title=f'{short}_DOWN')
    ws_dn.sheet_properties.tabColor = 'C2D4F4'
    write_venn_sheet(ws_dn, label, 'down',
                     aw_dn_df, i_dn_df, pa_dn_df,
                     'AW Decreased', 'Intox. Decreased', 'PA Decreased')

    print(f'{label}: UP={len(aw_up_df)} DOWN={len(aw_dn_df)}')

wb.save(OUT)
print(f'\nSaved: {OUT}')
