"""
Translocation analysis v2 — two-layer approach
------------------------------------------------
Layer 1 (biological filter): protein must be significant in BOTH fractions
  during AW vs naive AND in opposite directions (up in one, down in the other).

Layer 2 (statistical test): paired interaction t-test using per-animal
  (Chromatin − SN) balance. Tests whether that balance shifts between
  naive and AW, using the same-animal pairing across fractions.

Notes:
  - AW-M-3 excluded (global outlier)
  - Chromatin: Keep + Review rows only
  - Soluble Nuclear: all rows (no filter)
  - AW M3 excluded from AW replicates
"""

import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_pdf import PdfPages

FILE     = '../../EDIT  Philipp Alcohol proteome Simplified - Jan 2024  copy.xlsx'
OUT_XLSX = 'Translocation_v2_stats.xlsx'
OUT_PDF  = 'Translocation_v2_plots.pdf'

CORR_THRESH = 3.3
FC_THRESH   = 0.5

# AW-M-3 excluded
NAIVE_SUFFIXES = ['N-F-1', 'N-F-2', 'N-F-3', 'N-M-1', 'N-M-2']
AW_SUFFIXES    = ['AW-F-1', 'AW-F-2', 'AW-M-1', 'AW-M-2']   # AW-M-3 excluded

# Animal labels for pairing (same order as suffixes above)
NAIVE_LABELS = NAIVE_SUFFIXES
AW_LABELS    = AW_SUFFIXES


# ── Stats helpers ──────────────────────────────────────────────────────────────

def get_cols(df, prefix, suffixes):
    return [c for c in df.columns if any(prefix + s in str(c) for s in suffixes)]


def calc_aw_stats(df, naive_cols, aw_cols):
    """Return fold change and BH-corrected -log2(p) for AW vs naive."""
    naive = df[naive_cols].apply(pd.to_numeric, errors='coerce')
    cond  = df[aw_cols].apply(pd.to_numeric, errors='coerce')
    fc    = cond.mean(axis=1) - naive.mean(axis=1)
    pvs   = []
    for i in range(len(df)):
        n = naive.iloc[i].dropna().values
        c = cond.iloc[i].dropna().values
        pvs.append(stats.ttest_ind(n, c, equal_var=False)[1]
                   if len(n) >= 2 and len(c) >= 2 else np.nan)
    pvs   = pd.Series(pvs, index=df.index)
    valid = pvs.notna()
    ranks = pvs[valid].rank(ascending=False)
    corr  = pd.Series(np.nan, index=df.index)
    corr[valid] = -np.log2(np.minimum(1, pvs[valid] * valid.sum() / ranks))
    return fc, corr


# ── Load compartments ──────────────────────────────────────────────────────────

# Chromatin: Keep + Review only
chrom_raw = pd.read_excel(FILE, sheet_name='Chromatin', header=0)
chrom_df  = chrom_raw[chrom_raw['Filter'].isin(['Keep', 'Review'])].reset_index(drop=True)
chrom_naive_cols = get_cols(chrom_df, 'Chrom_', NAIVE_SUFFIXES)
chrom_aw_cols    = get_cols(chrom_df, 'Chrom_', AW_SUFFIXES)

# Soluble Nuclear: all rows
nuc_raw  = pd.read_excel(FILE, sheet_name='Soluble nuclear', header=0)
nuc_df   = nuc_raw.copy().reset_index(drop=True)
nuc_naive_cols = get_cols(nuc_df, 'Nuc_', NAIVE_SUFFIXES)
nuc_aw_cols    = get_cols(nuc_df, 'Nuc_', AW_SUFFIXES)

print(f'Chromatin proteins (Keep+Review): {len(chrom_df)}')
print(f'Soluble Nuclear proteins (all):   {len(nuc_df)}')
print(f'Chromatin naive cols: {chrom_naive_cols}')
print(f'Chromatin AW cols:    {chrom_aw_cols}')
print(f'Nuc naive cols:       {nuc_naive_cols}')
print(f'Nuc AW cols:          {nuc_aw_cols}')

# Per-fraction AW stats
chrom_fc, chrom_corr = calc_aw_stats(chrom_df, chrom_naive_cols, chrom_aw_cols)
nuc_fc,   nuc_corr   = calc_aw_stats(nuc_df,   nuc_naive_cols,   nuc_aw_cols)

# Gene → stats lookup
chrom_stats = pd.DataFrame({
    'FC':   chrom_fc.values,
    'Corr': chrom_corr.values,
}, index=chrom_df['Gene symbol'].astype(str))

nuc_stats = pd.DataFrame({
    'FC':   nuc_fc.values,
    'Corr': nuc_corr.values,
}, index=nuc_df['Gene symbol'].astype(str))

# Deduplicate (keep first occurrence)
chrom_stats = chrom_stats[~chrom_stats.index.duplicated(keep='first')]
nuc_stats   = nuc_stats[~nuc_stats.index.duplicated(keep='first')]

# Gene → raw value lookup (for per-animal paired test)
chrom_vals = chrom_df.drop_duplicates('Gene symbol').set_index(
    chrom_df.drop_duplicates('Gene symbol')['Gene symbol'].astype(str))
nuc_vals   = nuc_df.drop_duplicates('Gene symbol').set_index(
    nuc_df.drop_duplicates('Gene symbol')['Gene symbol'].astype(str))


# ── Layer 1: biological filter ─────────────────────────────────────────────────
# Significant in BOTH fractions AND opposite directions

def is_sig(stats_df, gene, direction):
    if gene not in stats_df.index:
        return False
    fc   = stats_df.loc[gene, 'FC']
    corr = stats_df.loc[gene, 'Corr']
    if not (np.isfinite(fc) and np.isfinite(corr)):
        return False
    if corr <= CORR_THRESH:
        return False
    if direction == 'up':
        return fc > FC_THRESH
    else:
        return fc < -FC_THRESH

# Union of all proteins detected in both compartments
all_chrom_genes = set(chrom_stats.index)
all_nuc_genes   = set(nuc_stats.index)
both_genes      = all_chrom_genes & all_nuc_genes

# Direction A: UP in Chromatin, DOWN in SN (protein moves INTO chromatin)
into_chrom_genes = {g for g in both_genes
                    if is_sig(chrom_stats, g, 'up') and is_sig(nuc_stats, g, 'down')}

# Direction B: DOWN in Chromatin, UP in SN (protein moves OUT of chromatin)
into_nuc_genes   = {g for g in both_genes
                    if is_sig(chrom_stats, g, 'down') and is_sig(nuc_stats, g, 'up')}

candidates = sorted(into_chrom_genes | into_nuc_genes)
print(f'\nLayer 1 candidates:')
print(f'  Into Chromatin (Chrom UP + SN DOWN): {len(into_chrom_genes)}')
print(f'  Into Nuclear   (Chrom DOWN + SN UP): {len(into_nuc_genes)}')
print(f'  Total candidates:                    {len(candidates)}')


# ── Layer 2: paired interaction t-test ────────────────────────────────────────
# For each animal compute (Chrom − SN) balance.
# Compare that balance between naive and AW animals.
# Uses the same-animal pairing across fractions.

def get_animal_vals(lookup_df, gene, cols):
    if gene not in lookup_df.index:
        return np.full(len(cols), np.nan)
    row = lookup_df.loc[gene]
    return np.array([float(pd.to_numeric(row[c], errors='coerce'))
                     if c in row.index else np.nan for c in cols])


rows = []
for gene in candidates:
    # Per-animal values for each fraction
    chrom_n = get_animal_vals(chrom_vals, gene, chrom_naive_cols)
    chrom_a = get_animal_vals(chrom_vals, gene, chrom_aw_cols)
    nuc_n   = get_animal_vals(nuc_vals,   gene, nuc_naive_cols)
    nuc_a   = get_animal_vals(nuc_vals,   gene, nuc_aw_cols)

    # Paired balance: Chrom − SN within each animal
    # Naive animals (5): match by position (same suffix order)
    balance_naive = chrom_n - nuc_n   # shape (5,)
    # AW animals (4)
    balance_aw    = chrom_a - nuc_a   # shape (4,)

    # Drop animals where either fraction is missing
    naive_ok = balance_naive[np.isfinite(balance_naive)]
    aw_ok    = balance_aw[np.isfinite(balance_aw)]

    if len(naive_ok) >= 2 and len(aw_ok) >= 2:
        _, p_interaction = stats.ttest_ind(aw_ok, naive_ok, equal_var=False)
        mean_balance_naive = float(np.mean(naive_ok))
        mean_balance_aw    = float(np.mean(aw_ok))
        delta_balance      = mean_balance_aw - mean_balance_naive
    else:
        p_interaction      = np.nan
        mean_balance_naive = np.nan
        mean_balance_aw    = np.nan
        delta_balance      = np.nan

    direction = 'Into_Chromatin' if gene in into_chrom_genes else 'Into_Nuclear'

    rows.append({
        'Gene':                gene,
        'Direction':           direction,
        'Chrom_AW_FC':         round(float(chrom_stats.loc[gene, 'FC']),   4),
        'Chrom_AW_corrected':  round(float(chrom_stats.loc[gene, 'Corr']), 4),
        'SN_AW_FC':            round(float(nuc_stats.loc[gene, 'FC']),     4),
        'SN_AW_corrected':     round(float(nuc_stats.loc[gene, 'Corr']),   4),
        'Mean_balance_naive':  round(mean_balance_naive, 4) if np.isfinite(mean_balance_naive) else np.nan,
        'Mean_balance_AW':     round(mean_balance_aw,    4) if np.isfinite(mean_balance_aw)    else np.nan,
        'Delta_balance':       round(delta_balance,      4) if np.isfinite(delta_balance)      else np.nan,
        'p_interaction_raw':   p_interaction,
    })

df_results = pd.DataFrame(rows)

# BH correction on interaction p-values
valid = df_results['p_interaction_raw'].notna()
if valid.sum() > 0:
    pv    = df_results.loc[valid, 'p_interaction_raw'].values
    order = np.argsort(pv)
    ranks = np.empty_like(order)
    ranks[order] = np.arange(1, len(pv) + 1)
    bh = np.minimum(1, pv * len(pv) / ranks)
    for i in range(len(bh) - 2, -1, -1):
        bh[i] = min(bh[i], bh[i + 1])
    df_results.loc[valid, 'p_interaction_BH'] = bh

df_results = df_results.sort_values(
    ['Direction', 'p_interaction_BH'], ascending=[True, True]
).reset_index(drop=True)

sig_results = df_results[df_results['p_interaction_BH'] < 0.05]
print(f'\nLayer 2 results (BH < 0.05):')
print(f'  Into Chromatin: {(sig_results["Direction"] == "Into_Chromatin").sum()}')
print(f'  Into Nuclear:   {(sig_results["Direction"] == "Into_Nuclear").sum()}')
print(f'  Total:          {len(sig_results)}')


# ── Save Excel ─────────────────────────────────────────────────────────────────

thin   = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

FILLS = {
    'Into_Chromatin': PatternFill('solid', fgColor='FCE4D6'),
    'Into_Nuclear':   PatternFill('solid', fgColor='DDEEFF'),
}
HDR_FILLS = {
    'Into_Chromatin': PatternFill('solid', fgColor='C55A11'),
    'Into_Nuclear':   PatternFill('solid', fgColor='2E75B6'),
}

def write_sheet(ws, df, title_fill):
    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill      = title_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border    = border

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name='Arial', size=9)
            cell.alignment = Alignment(horizontal='center')
            cell.border    = border
            # Highlight BH-significant rows
            col_name = headers[ci - 1]
            if col_name == 'p_interaction_BH' and isinstance(val, float) and val < 0.05:
                cell.fill = PatternFill('solid', fgColor='E2EFDA')

    for ci, h in enumerate(headers, 1):
        width = 35 if 'Gene' in h else 18
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

wb = Workbook()
wb.remove(wb.active)

# Sheet 1: all candidates (both directions)
ws_all = wb.create_sheet('All_candidates')
write_sheet(ws_all, df_results, PatternFill('solid', fgColor='1F4E79'))

# Sheet 2: Into Chromatin only
ws_ic = wb.create_sheet('Into_Chromatin')
write_sheet(ws_ic,
            df_results[df_results['Direction'] == 'Into_Chromatin'].reset_index(drop=True),
            HDR_FILLS['Into_Chromatin'])

# Sheet 3: Into Nuclear only
ws_in = wb.create_sheet('Into_Nuclear')
write_sheet(ws_in,
            df_results[df_results['Direction'] == 'Into_Nuclear'].reset_index(drop=True),
            HDR_FILLS['Into_Nuclear'])

# Sheet 4: BH-significant only
ws_sig = wb.create_sheet('Significant_BH05')
write_sheet(ws_sig, sig_results.reset_index(drop=True),
            PatternFill('solid', fgColor='375623'))

wb.save(OUT_XLSX)
print(f'\nSaved: {OUT_XLSX}')


# ── Plots ──────────────────────────────────────────────────────────────────────

with PdfPages(OUT_PDF) as pdf:

    # 1. Summary bar: how many candidates at each stage
    fig, ax = plt.subplots(figsize=(7, 4))
    fig.patch.set_facecolor('white')
    stages = ['Layer 1\n(opposite\ndirections)',
              'Layer 2\n(BH < 0.05)']
    into_c = [len(into_chrom_genes),
              (sig_results['Direction'] == 'Into_Chromatin').sum()]
    into_n = [len(into_nuc_genes),
              (sig_results['Direction'] == 'Into_Nuclear').sum()]
    x = np.arange(len(stages))
    w = 0.35
    ax.bar(x - w/2, into_c, w, color='#E8305A', label='Into Chromatin')
    ax.bar(x + w/2, into_n, w, color='#2B7FD4', label='Into Nuclear')
    for i, (c, n) in enumerate(zip(into_c, into_n)):
        ax.text(i - w/2, c + 0.3, str(c), ha='center', va='bottom', fontsize=11, fontweight='bold')
        ax.text(i + w/2, n + 0.3, str(n), ha='center', va='bottom', fontsize=11, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(stages, fontsize=11)
    ax.set_ylabel('Number of proteins', fontsize=11)
    ax.set_title('Translocation candidates — two-layer filter\n(AW vs Naïve, AW-M-3 excluded)',
                 fontsize=12, fontweight='bold')
    ax.legend(fontsize=10)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    plt.tight_layout()
    pdf.savefig(fig, dpi=150, bbox_inches='tight')
    plt.close(fig)

    # 2. Scatter: Chrom FC vs SN FC, colored by direction
    fig, ax = plt.subplots(figsize=(7, 6))
    fig.patch.set_facecolor('white')

    chrom_fc_all = df_results['Chrom_AW_FC'].values
    sn_fc_all    = df_results['SN_AW_FC'].values
    sig_mask     = df_results['p_interaction_BH'].fillna(1) < 0.05
    dir_mask_ic  = df_results['Direction'] == 'Into_Chromatin'
    dir_mask_in  = df_results['Direction'] == 'Into_Nuclear'

    # All candidates (not sig) — grey
    ax.scatter(sn_fc_all[~sig_mask], chrom_fc_all[~sig_mask],
               s=30, color='#CCCCCC', alpha=0.6, linewidths=0, label='Not sig. (BH≥0.05)')
    # Sig into chromatin
    ax.scatter(sn_fc_all[sig_mask & dir_mask_ic], chrom_fc_all[sig_mask & dir_mask_ic],
               s=60, color='#E8305A', alpha=0.9, linewidths=0,
               label=f'Into Chromatin, sig. (n={(sig_mask & dir_mask_ic).sum()})')
    # Sig into nuclear
    ax.scatter(sn_fc_all[sig_mask & dir_mask_in], chrom_fc_all[sig_mask & dir_mask_in],
               s=60, color='#2B7FD4', alpha=0.9, linewidths=0,
               label=f'Into Nuclear, sig. (n={(sig_mask & dir_mask_in).sum()})')

    ax.axhline(0, color='#AAAAAA', linewidth=0.8)
    ax.axvline(0, color='#AAAAAA', linewidth=0.8)
    ax.axhline( FC_THRESH, color='#888888', linestyle='--', linewidth=1.0, alpha=0.5)
    ax.axhline(-FC_THRESH, color='#888888', linestyle='--', linewidth=1.0, alpha=0.5)
    ax.axvline( FC_THRESH, color='#888888', linestyle='--', linewidth=1.0, alpha=0.5)
    ax.axvline(-FC_THRESH, color='#888888', linestyle='--', linewidth=1.0, alpha=0.5)

    ax.set_xlabel('SN fold change  (AW − Naïve)', fontsize=12)
    ax.set_ylabel('Chromatin fold change  (AW − Naïve)', fontsize=12)
    ax.set_title('Translocation candidates\n(significant in both fractions, opposite directions)',
                 fontsize=11, fontweight='bold')
    ax.legend(fontsize=9, framealpha=0.9)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    # Annotate quadrants
    xlim = ax.get_xlim(); ylim = ax.get_ylim()
    ax.text(xlim[0] + 0.1, ylim[1] - 0.1, 'SN↓ / Chrom↑\n(into Chromatin)',
            fontsize=8, color='#E8305A', va='top')
    ax.text(xlim[1] - 0.1, ylim[0] + 0.1, 'SN↑ / Chrom↓\n(into Nuclear)',
            fontsize=8, color='#2B7FD4', ha='right')
    plt.tight_layout()
    pdf.savefig(fig, dpi=150, bbox_inches='tight')
    plt.close(fig)

    # 3. Delta balance plot for significant proteins
    if len(sig_results) > 0:
        for direction, color, label in [
            ('Into_Chromatin', '#E8305A', 'Into Chromatin'),
            ('Into_Nuclear',   '#2B7FD4', 'Into Nuclear'),
        ]:
            sub = sig_results[sig_results['Direction'] == direction].copy()
            if len(sub) == 0:
                continue
            sub = sub.sort_values('Delta_balance', ascending=(direction == 'Into_Nuclear'))
            fig, ax = plt.subplots(figsize=(max(6, len(sub) * 0.4 + 2), 5))
            fig.patch.set_facecolor('white')
            x = np.arange(len(sub))
            ax.bar(x, sub['Delta_balance'].values, color=color, alpha=0.8, edgecolor='white')
            ax.axhline(0, color='#333333', linewidth=1.0)
            ax.set_xticks(x)
            ax.set_xticklabels(sub['Gene'].values, rotation=45, ha='right', fontsize=8)
            ax.set_ylabel('Δ balance  (Chrom−SN)_AW − (Chrom−SN)_naive', fontsize=10)
            ax.set_title(f'{label} — significant proteins (BH < 0.05)\n'
                         f'Δ balance = shift in Chromatin/SN ratio during AW',
                         fontsize=11, fontweight='bold')
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            plt.tight_layout()
            pdf.savefig(fig, dpi=150, bbox_inches='tight')
            plt.close(fig)

print(f'Saved: {OUT_PDF}')
