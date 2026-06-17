import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE = '../../EDIT  Philipp Alcohol proteome Simplified - Jan 2024  copy.xlsx'
OUT  = 'Chrom_vs_SN_Balance.xlsx'

CORR_THRESH = 3.3
FC_THRESH   = 0.5

AW_SUFFIXES = ['AW-F-1', 'AW-F-2', 'AW-M-1', 'AW-M-2']

CONDITIONS = {
    'Naive':    {'chrom_pat': 'Chrom_N-',  'nuc_pat': 'Nuc_N-',  'suffixes': None},
    'Intox':    {'chrom_pat': 'Chrom_I-',  'nuc_pat': 'Nuc_I-',  'suffixes': None},
    'AW':       {'chrom_pat': 'Chrom_AW-', 'nuc_pat': 'Nuc_AW-', 'suffixes': AW_SUFFIXES},
    'PA':       {'chrom_pat': 'Chrom_PA-', 'nuc_pat': 'Nuc_PA-', 'suffixes': None},
}

# Load sheets
chrom_raw = pd.read_excel(FILE, sheet_name='Chromatin', header=0)
nuc_raw   = pd.read_excel(FILE, sheet_name='Soluble nuclear', header=0)

# Filter Chromatin to Keep + Review
chrom_df = chrom_raw[chrom_raw['Filter'].isin(['Keep', 'Review'])].reset_index(drop=True)
nuc_df   = nuc_raw.copy().reset_index(drop=True)

# Merge on Accession — only proteins present in both
chrom_df['Accession'] = chrom_df['Accession'].astype(str).str.strip()
nuc_df['Accession']   = nuc_df['Accession'].astype(str).str.strip()

merged = pd.merge(
    chrom_df[['Accession', 'Gene symbol', 'Description', 'Filter']],
    nuc_df[['Accession']],
    on='Accession', how='inner'
).drop_duplicates(subset='Accession').reset_index(drop=True)

print(f'Proteins in both fractions (Chrom Keep/Review + SN): {len(merged)}')


def get_cols(df, pat, suffixes=None):
    if suffixes:
        return [c for c in df.columns if any(pat.rstrip('-') + '-' + s.split('-', 1)[1] in str(c)
                                             or pat + s.split('-', 1)[1] in str(c)
                                             for s in suffixes)]
    return [c for c in df.columns if pat in str(c)]


def get_cols_aw(df, pat_prefix, suffixes):
    """Match columns like Chrom_AW-F-1 etc. using explicit suffix list."""
    return [c for c in df.columns if any(pat_prefix + s in str(c) for s in suffixes)]


def calc_balance_stats(chrom_vals, nuc_vals):
    """FC = mean(Chrom) - mean(SN), Welch's t-test, BH correction across all proteins."""
    return chrom_vals.mean(axis=1) - nuc_vals.mean(axis=1)


def bh_correct(p_series):
    valid = p_series.notna()
    ranks = p_series[valid].rank(ascending=False)
    corrected = pd.Series(np.nan, index=p_series.index)
    corrected[valid] = -np.log2(np.minimum(1, p_series[valid] * valid.sum() / ranks))
    return corrected


# Compute stats for each condition
results = merged[['Accession', 'Gene symbol', 'Description', 'Filter']].copy()

for cond, cfg in CONDITIONS.items():
    chrom_pat = cfg['chrom_pat']
    nuc_pat   = cfg['nuc_pat']
    suffixes  = cfg['suffixes']

    if suffixes:
        # e.g. suffix='AW-F-1', chrom_pat='Chrom_AW-' → match 'Chrom_AW-F-1'
        chrom_cols = [c for c in chrom_df.columns if any(chrom_pat + s.split('-', 1)[1] in str(c) for s in suffixes)]
        nuc_cols   = [c for c in nuc_df.columns   if any(nuc_pat   + s.split('-', 1)[1] in str(c) for s in suffixes)]
    else:
        chrom_cols = [c for c in chrom_df.columns if chrom_pat in str(c)]
        nuc_cols   = [c for c in nuc_df.columns   if nuc_pat   in str(c)]

    print(f'{cond}: Chrom cols = {chrom_cols}')
    print(f'{cond}: SN cols    = {nuc_cols}')

    # Pull values aligned to merged accessions
    chrom_vals = chrom_df.set_index('Accession').loc[merged['Accession'], chrom_cols].apply(pd.to_numeric, errors='coerce')
    nuc_vals   = nuc_df.set_index('Accession').loc[merged['Accession'], nuc_cols].apply(pd.to_numeric, errors='coerce')
    chrom_vals.index = merged.index
    nuc_vals.index   = merged.index

    fc = chrom_vals.mean(axis=1) - nuc_vals.mean(axis=1)

    p_vals = []
    for i in range(len(merged)):
        cv = chrom_vals.iloc[i].dropna().values
        nv = nuc_vals.iloc[i].dropna().values
        if len(cv) >= 2 and len(nv) >= 2:
            _, p = stats.ttest_ind(cv, nv, equal_var=False)
            p_vals.append(p)
        else:
            p_vals.append(np.nan)

    p_series  = pd.Series(p_vals, index=merged.index)
    corrected = bh_correct(p_series)

    sig = pd.Series('NS', index=merged.index)
    sig[(corrected > CORR_THRESH) & (fc >  FC_THRESH)] = 'Up'    # more in Chrom
    sig[(corrected > CORR_THRESH) & (fc < -FC_THRESH)] = 'Down'  # more in SN

    results[f'FC_{cond}']        = fc.round(4)
    results[f'pval_{cond}']      = p_series.round(6)
    results[f'Corrected_{cond}'] = corrected.round(4)
    results[f'Sig_{cond}']       = sig

    up   = (sig == 'Up').sum()
    down = (sig == 'Down').sum()
    print(f'  {cond}: Up(Chrom>SN)={up}  Down(SN>Chrom)={down}')


# ── Write Excel ──
wb = Workbook()
ws = wb.active
ws.title = 'Chrom_vs_SN_Balance'

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COND_FILLS = {
    'Naive': PatternFill('solid', fgColor='D9D9D9'),
    'Intox': PatternFill('solid', fgColor='2E75B6'),
    'AW':    PatternFill('solid', fgColor='375623'),
    'PA':    PatternFill('solid', fgColor='7030A0'),
}
INFO_FILL = PatternFill('solid', fgColor='1F4E79')
UP_FILL   = PatternFill('solid', fgColor='FCE4D6')
DOWN_FILL = PatternFill('solid', fgColor='DDEEFF')
HDR_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
CELL_FONT = Font(name='Arial', size=9)
CENTER    = Alignment(horizontal='center', wrap_text=True)

headers = list(results.columns)
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font      = HDR_FONT
    cell.alignment = CENTER
    cond_match = next((c for c in COND_FILLS if f'_{c}' in h), None)
    cell.fill = COND_FILLS[cond_match] if cond_match else INFO_FILL

for row_idx, row in enumerate(results.itertuples(index=False), 2):
    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font   = CELL_FONT
        cell.border = border
        col_name = headers[col_idx - 1]
        if col_name.startswith('Sig_'):
            if val == 'Up':
                cell.fill = UP_FILL
                cell.font = Font(name='Arial', size=9, bold=True, color='C55A11')
            elif val == 'Down':
                cell.fill = DOWN_FILL
                cell.font = Font(name='Arial', size=9, bold=True, color='2E75B6')

ws.column_dimensions[get_column_letter(1)].width = 12
ws.column_dimensions[get_column_letter(2)].width = 12
ws.column_dimensions[get_column_letter(3)].width = 35
ws.column_dimensions[get_column_letter(4)].width = 8
for col_idx in range(5, len(headers) + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 14

ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

wb.save(OUT)
print(f'\nSaved: {OUT}')
