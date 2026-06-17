import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE = '../EDIT  Philipp Alcohol proteome Simplified - Jan 2024  copy.xlsx'
OUT  = 'Volcano_Plot_Data.xlsx'

CORR_THRESH = 3.3
FC_THRESH   = 0.5

SHEETS = [
    ('Membrane',        'Membrane',        'Memb_',  'Memb_N-',  'all'),
    ('Cytosol',         'Cytosol',         'Cyto_',  'Cyto_N-',  'all'),
    ('Chromatin',       'Chromatin',       'Chrom_', 'Chrom_N-', 'keep_review'),
    ('Soluble_nuclear', 'Soluble nuclear', 'Nuc_',   'Nuc_N-',   'all'),          # fixed: was keep_review
]

# Explicit replicate lists — AW-M-3 excluded from AW (global outlier)
COND_SUFFIXES = {
    'Intoxication':          ['I-F-1',  'I-F-2',  'I-F-3',  'I-M-1',  'I-M-2'],
    'Acute_Withdrawal':      ['AW-F-1', 'AW-F-2', 'AW-M-1', 'AW-M-2'],   # AW-M-3 excluded
    'Protracted_Abstinence': ['PA-F-1', 'PA-F-2', 'PA-F-3', 'PA-M-1', 'PA-M-2'],
}
CONDITIONS = list(COND_SUFFIXES.keys())

def calc_stats(df, naive_cols, cond_cols):
    naive = df[naive_cols].apply(pd.to_numeric, errors='coerce')
    cond  = df[cond_cols].apply(pd.to_numeric, errors='coerce')
    fc    = cond.mean(axis=1) - naive.mean(axis=1)
    p_vals = []
    for i in range(len(df)):
        n = naive.iloc[i].dropna().values
        c = cond.iloc[i].dropna().values
        if len(n) >= 2 and len(c) >= 2:
            _, p = stats.ttest_ind(n, c, equal_var=False)
            p_vals.append(p)
        else:
            p_vals.append(np.nan)
    p_vals    = pd.Series(p_vals, index=df.index)
    valid     = p_vals.notna()
    ranks     = p_vals[valid].rank(ascending=False)
    corrected = pd.Series(np.nan, index=df.index)
    corrected[valid] = -np.log2(np.minimum(1, p_vals[valid] * valid.sum() / ranks))
    return fc, corrected

wb = Workbook()
wb.remove(wb.active)

# Styles
HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
COND_FILLS = {
    'Intoxication':          PatternFill('solid', fgColor='2E75B6'),
    'Acute_Withdrawal':      PatternFill('solid', fgColor='375623'),
    'Protracted_Abstinence': PatternFill('solid', fgColor='7030A0'),
}
UP_FILL   = PatternFill('solid', fgColor='FCE4D6')
DOWN_FILL = PatternFill('solid', fgColor='DDEEFF')
HDR_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
CELL_FONT = Font(name='Arial', size=9)
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for sheet_key, excel_name, prefix, naive_pat, fmode in SHEETS:
    raw = pd.read_excel(FILE, sheet_name=excel_name, header=0)

    if fmode == 'keep_review':
        df = raw[raw['Filter'].isin(['Keep', 'Review'])].reset_index(drop=True)
    else:
        df = raw.copy().reset_index(drop=True)

    # Build result dataframe
    info_cols = ['Accession', 'Gene symbol', 'Description']
    if 'Filter' in df.columns:
        info_cols.append('Filter')
    result = df[info_cols].copy()

    for cond_label in CONDITIONS:
        suffixes   = COND_SUFFIXES[cond_label]
        naive_cols = [c for c in df.columns if naive_pat in str(c)]
        cond_cols  = [c for c in df.columns if any(prefix + s in str(c) for s in suffixes)]
        fc, corrected = calc_stats(df, naive_cols, cond_cols)

        sig = pd.Series('NS', index=df.index)
        sig[(corrected > CORR_THRESH) & (fc >  FC_THRESH)] = 'Up'
        sig[(corrected > CORR_THRESH) & (fc < -FC_THRESH)] = 'Down'

        result[f'FC_{cond_label}']        = fc.round(4)
        result[f'Corrected_{cond_label}'] = corrected.round(4)
        result[f'Sig_{cond_label}']       = sig

    ws = wb.create_sheet(title=sheet_key[:31])

    # Header row
    headers = list(result.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        # colour by condition group
        cond_match = next((c for c in COND_FILLS if c in h), None)
        if cond_match:
            cell.fill = COND_FILLS[cond_match]
        else:
            cell.fill = HDR_FILL

    # Data rows
    for row_idx, row in enumerate(result.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = CELL_FONT
            cell.border = border
            # Highlight sig cells
            col_name = headers[col_idx - 1]
            if col_name.startswith('Sig_'):
                if val == 'Up':
                    cell.fill = UP_FILL
                    cell.font = Font(name='Arial', size=9, bold=True, color='C55A11')
                elif val == 'Down':
                    cell.fill = DOWN_FILL
                    cell.font = Font(name='Arial', size=9, bold=True, color='2E75B6')

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 12  # Accession
    ws.column_dimensions[get_column_letter(2)].width = 12  # Gene symbol
    ws.column_dimensions[get_column_letter(3)].width = 35  # Description
    for col_idx in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    print(f'  {sheet_key}: {len(result)} proteins')

wb.save(OUT)
print(f'\nSaved: {OUT}')
