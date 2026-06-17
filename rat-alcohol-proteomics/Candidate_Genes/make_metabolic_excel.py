"""
Metabolic gene table → Excel
=============================
All 111 genes × 4 fractions × 3 conditions.
Columns: Gene | Membrane FC/Corr (Intox/AW/PA) | Cytosol ... | Chromatin ... | SN ...
Cells: red=UP, blue=DOWN, grey=not detected, white=ns.

Output: MetabolicGenes_Table.xlsx
"""

import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE  = os.path.join(SCRIPT_DIR, '..',
             'EDIT  Excluding AWM3 Philipp Alcohol proteome Simplified - Jan 2024  copy.xlsx')
OUT_XLSX   = os.path.join(SCRIPT_DIR, 'MetabolicGenes_Table.xlsx')

FC_THRESH   = 0.5
CORR_THRESH = 3.3

GENE_LIST = [
    ('ACSS2',           'Acss2'),
    ('ACSS1',           'Acss1'),
    ('ACLY',            'Acly'),
    ('ACACA',           'Acaca'),
    ('SLC16A1 (MCT1)',  'Slc16a1'),
    ('SLC16A7 (MCT2)',  'Slc16a7'),
    ('SLC16A3 (MCT4)',  'Slc16a3'),
    ('CS',              'Cs'),
    ('ACO2',            'Aco2'),
    ('IDH2',            'Idh2'),
    ('DLAT',            'Dlat'),
    ('PDHA1',           'Pdha1'),
    ('SUCLG1',          'Suclg1'),
    ('SDHA',            'Sdha'),
    ('SLC2A1 (GLUT1)',  'Slc2a1'),
    ('SLC2A3 (GLUT3)',  'Slc2a3'),
    ('HK1',             'Hk1'),
    ('PFKM',            'Pfkm'),
    ('ALDOA',           'Aldoa'),
    ('GAPDH',           'Gapdh'),
    ('PKM',             'Pkm'),
    ('LDHA',            'Ldha'),
    ('LDHB',            'Ldhb'),
    ('ADH1',            'Adh1'),
    ('CYP2E1',          'Cyp2e1'),
    ('CAT',             'Cat'),
    ('ADH5',            'Adh5'),
    ('ALDH2',           'Aldh2'),
    ('ALDH1A1',         'Aldh1a1'),
    ('SIRT1',           'Sirt1'),
    ('SIRT3',           'Sirt3'),
    ('NAMPT',           'Nampt'),
    ('PARP1',           'Parp1'),
    ('NFE2L2 (NRF2)',   'Nfe2l2'),
    ('SOD2',            'Sod2'),
    ('GPX4',            'Gpx4'),
    ('PRDX3',           'Prdx3'),
    ('TXN',             'Txn'),
    ('GCLC',            'Gclc'),
    ('EP300 (p300)',     'Ep300'),
    ('CREBBP (CBP)',     'Crebbp'),
    ('KAT2A',           'Kat2a'),
    ('HDAC1',           'Hdac1'),
    ('HDAC2',           'Hdac2'),
    ('SIRT6',           'Sirt6'),
    ('GFAP',            'Gfap'),
    ('GLUL',            'Glul'),
    ('SLC1A2 (EAAT2)',  'Slc1a2'),
    ('SLC1A3 (EAAT1)',  'Slc1a3'),
    ('GOT1',            'Got1'),
    ('SHMT1',           'Shmt1'),
    ('DLD',             'Dld'),
    ('FH',              'Fh'),
    ('MDH2',            'Mdh2'),
    ('MDH1',            'Mdh1'),
    ('GOT2',            'Got2'),
    ('OGDH',            'Ogdh'),
    ('ATP5A1',          'Atp5a1'),
    ('NDUFS1',          'Ndufs1'),
    ('UQCRC1',          'Uqcrc1'),
    ('COX4I1',          'Cox4i1'),
    ('TFAM',            'Tfam'),
    ('PPARGC1A (PGC1a)','Ppargc1a'),
    ('HIF1A',           'Hif1a'),
    ('CREB1',           'Creb1'),
    ('CAMK2A',          'Camk2a'),
    ('CAMK2B',          'Camk2b'),
    ('PRKAA1 (AMPK)',   'Prkaa1'),
    ('PRKAA2 (AMPK)',   'Prkaa2'),
    ('MTOR',            'Mtor'),
    ('RPTOR',           'Rptor'),
    ('RICTOR',          'Rictor'),
    ('FOXO3',           'Foxo3'),
    ('HSPA9',           'Hspa9'),
    ('VDAC1',           'Vdac1'),
    ('TOMM20',          'Tomm20'),
    ('CPT1A',           'Cpt1a'),
    ('CPT2',            'Cpt2'),
    ('ACADM',           'Acadm'),
    ('HADHA',           'Hadha'),
    ('HADHB',           'Hadhb'),
    ('G6PD',            'G6pd'),
    ('PGD',             'Pgd'),
    ('ME1',             'Me1'),
    ('ME2',             'Me2'),
    ('GLUD1',           'Glud1'),
    ('GLS',             'Gls'),
    ('PC',              'Pc'),
    ('PDC',             'Pdc'),
    ('PDK1',            'Pdk1'),
    ('PDK2',            'Pdk2'),
    ('FASN',            'Fasn'),
    ('SREBF1',          'Srebf1'),
    ('HMOX1',           'Hmox1'),
    ('NOS2',            'Nos2'),
    ('ARG1',            'Arg1'),
    ('IL1B',            'Il1b'),
    ('TNF',             'Tnf'),
    ('RELA',            'Rela'),
    ('NFKB1',           'Nfkb1'),
    ('ATF4',            'Atf4'),
    ('DDIT3 (CHOP)',    'Ddit3'),
    ('XBP1',            'Xbp1'),
    ('EIF2AK3 (PERK)',  'Eif2ak3'),
    ('IRE1 (ERN1)',      'Ern1'),
    ('ATG5',            'Atg5'),
    ('ATG7',            'Atg7'),
    ('MAP1LC3B (LC3B)', 'Map1lc3b'),
    ('SQSTM1 (p62)',    'Sqstm1'),
    ('BNIP3',           'Bnip3'),
    ('PINK1',           'Pink1'),
    ('PRKN (Parkin)',   'Prkn'),
    # PDC subunits
    ('PDHB',            'Pdhb'),
    ('PDHX',            'Pdhx'),
]

FRACTIONS = [
    ('Membrane',        'Memb'),
    ('Cytosol',         'Cyto'),
    ('Chromatin',       'Chrom'),
    ('Soluble nuclear', 'SN'),
]

COND_COLS = {
    'Intox': ('Fold change',   'Corrected'),
    'AW':    ('Fold change.1', 'Corrected.1'),
    'PA':    ('Fold change.2', 'Corrected.2'),
}
CONDITIONS = ['Intox', 'AW', 'PA']

# ── Styles ─────────────────────────────────────────────────────────────────────
FILL_UP   = PatternFill(fill_type='solid', fgColor='FADADD')
FILL_DOWN = PatternFill(fill_type='solid', fgColor='DAE8FA')
FILL_MISS = PatternFill(fill_type='solid', fgColor='F0F0F0')
FILL_NS   = PatternFill(fill_type='solid', fgColor='FFFFFF')
FILL_HEAD = PatternFill(fill_type='solid', fgColor='D9D9D9')
FILL_FRAC = PatternFill(fill_type='solid', fgColor='EEF0F8')

THIN   = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left',   vertical='center')

def cell_style(ws, row, col, value, fill, font, align=CENTER):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill; c.font = font; c.alignment = align; c.border = BORDER
    return c

# ── Load sheets ────────────────────────────────────────────────────────────────
sheets = {}
for frac_name, _ in FRACTIONS:
    df = pd.read_excel(DATA_FILE, sheet_name=frac_name)
    df['Gene symbol'] = df['Gene symbol'].astype(str).str.strip()
    if 'Filter' in df.columns:
        df = df[df['Filter'].isin(['Keep', 'Review'])]
    sheets[frac_name] = df

# ── Build workbook ─────────────────────────────────────────────────────────────
wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = 'Metabolic Genes'

# Row 1: Gene | then per fraction: merged 6-col header
# Row 2: (blank for gene) | then per cond: FC, Corr  ×3
# Each fraction = 6 columns (Intox FC, Intox Corr, AW FC, AW Corr, PA FC, PA Corr)

GENE_COL  = 1
DATA_START = 2   # first data column

frac_col_starts = {}
col = DATA_START
for frac_name, _ in FRACTIONS:
    frac_col_starts[frac_name] = col
    col += 6  # 3 conditions × 2 values

TOTAL_COLS = col - 1

# ── Row 1: title headers ───────────────────────────────────────────────────────
ws.merge_cells(start_row=1, start_column=GENE_COL, end_row=2, end_column=GENE_COL)
cell_style(ws, 1, GENE_COL, 'Gene', FILL_HEAD,
           Font(bold=True, name='Arial', size=11))

for frac_name, _ in FRACTIONS:
    fc = frac_col_starts[frac_name]
    ws.merge_cells(start_row=1, start_column=fc, end_row=1, end_column=fc + 5)
    cell_style(ws, 1, fc, frac_name, FILL_FRAC,
               Font(bold=True, name='Arial', size=11))

# ── Row 2: condition + FC/Corr subheaders ─────────────────────────────────────
for frac_name, _ in FRACTIONS:
    fc = frac_col_starts[frac_name]
    for i, cond in enumerate(CONDITIONS):
        c_fc   = fc + i * 2
        c_corr = c_fc + 1
        cell_style(ws, 2, c_fc,   f'{cond} FC',     FILL_HEAD,
                   Font(bold=True, name='Arial', size=10))
        cell_style(ws, 2, c_corr, f'{cond} Corr p', FILL_HEAD,
                   Font(bold=True, name='Arial', size=10))

# ── Data rows ──────────────────────────────────────────────────────────────────
for gene_i, (display, rat_sym) in enumerate(GENE_LIST):
    row = gene_i + 3

    # Gene name
    cell_style(ws, row, GENE_COL, display, FILL_NS,
               Font(bold=True, name='Arial', size=10), LEFT)

    for frac_name, _ in FRACTIONS:
        df    = sheets[frac_name]
        match = df[df['Gene symbol'].str.lower() == rat_sym.lower()]
        fc0   = frac_col_starts[frac_name]

        for i, cond in enumerate(CONDITIONS):
            c_fc   = fc0 + i * 2
            c_corr = c_fc + 1

            if match.empty:
                for col in (c_fc, c_corr):
                    cell_style(ws, row, col, '—', FILL_MISS,
                               Font(name='Arial', size=10, color='AAAAAA'))
            else:
                r        = match.iloc[0]
                fc_col, corr_col = COND_COLS[cond]
                fc_val   = pd.to_numeric(r.get(fc_col,   np.nan), errors='coerce')
                corr_val = pd.to_numeric(r.get(corr_col, np.nan), errors='coerce')

                sig = (not pd.isna(fc_val) and not pd.isna(corr_val)
                       and corr_val > CORR_THRESH and abs(fc_val) > FC_THRESH)
                if sig and fc_val > 0:
                    fill = FILL_UP
                    font = Font(name='Arial', size=10, bold=True)
                elif sig and fc_val < 0:
                    fill = FILL_DOWN
                    font = Font(name='Arial', size=10, bold=True)
                else:
                    fill = FILL_NS
                    font = Font(name='Arial', size=10, color='555555')

                fc_str   = f'{fc_val:+.2f}'  if not pd.isna(fc_val)   else '—'
                corr_str = f'{corr_val:.1f}' if not pd.isna(corr_val) else '—'
                cell_style(ws, row, c_fc,   fc_str,   fill, font)
                cell_style(ws, row, c_corr, corr_str, fill, font)

# ── Column widths ──────────────────────────────────────────────────────────────
ws.column_dimensions[get_column_letter(GENE_COL)].width = 20
for frac_name, _ in FRACTIONS:
    fc = frac_col_starts[frac_name]
    for i in range(6):
        ws.column_dimensions[get_column_letter(fc + i)].width = 10

ws.row_dimensions[1].height = 18
ws.row_dimensions[2].height = 18
ws.freeze_panes = 'B3'

# ── Legend sheet ───────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Legend')
legend_rows = [
    ('Color',       'Meaning'),
    ('Red (bold)',  'Upregulated: FC > +0.5 AND corr. −log2 p > 3.3'),
    ('Blue (bold)', 'Downregulated: FC < −0.5 AND corr. −log2 p > 3.3'),
    ('Grey',        'Not detected in this fraction'),
    ('White',       'Detected but not significant'),
    ('',            ''),
    ('FC',          'Log2 fold change vs Naïve'),
    ('Corr p',      '−log2 BH-corrected p-value'),
    ('',            ''),
    ('Notes:',      'AW-M-3 excluded from all AW calculations'),
    ('',            'Chromatin: Keep+Review rows only'),
]
for r, (a, b) in enumerate(legend_rows, 1):
    ws2.cell(row=r, column=1, value=a).font = Font(bold=True, name='Arial', size=10)
    ws2.cell(row=r, column=2, value=b).font = Font(name='Arial', size=10)
ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 55

wb.save(OUT_XLSX)
print(f'Saved: {OUT_XLSX}')
print(f'Genes: {len(GENE_LIST)}')
