"""
PDC components Excel file
=========================
Pyruvate dehydrogenase complex subunits:
  PDHA1 (Pdha1), PDHB (Pdhb), DLAT (Dlat), DLD (Dld), PDHX (Pdhx)

Checks all 4 fractions × 3 conditions.
Cells highlighted: red = UP, blue = DOWN (|FC|>0.5 AND corr>3.3).
Grey = not detected in that fraction.

Output: PDC_Components.xlsx
"""

import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE  = os.path.join(SCRIPT_DIR, '..',
             'EDIT  Excluding AWM3 Philipp Alcohol proteome Simplified - Jan 2024  copy.xlsx')
OUT_XLSX   = os.path.join(SCRIPT_DIR, 'PDC_Components.xlsx')

FC_THRESH   = 0.5
CORR_THRESH = 3.3

PDC_GENES = [
    ('PDHA1', 'Pdha1'),
    ('PDHB',  'Pdhb'),
    ('DLAT',  'Dlat'),
    ('DLD',   'Dld'),
    ('PDHX',  'Pdhx'),
]

FRACTIONS = [
    ('Membrane',       'Memb'),
    ('Cytosol',        'Cyto'),
    ('Chromatin',      'Chrom'),
    ('Soluble nuclear','SN'),
]

COND_COLS = {
    'Intox': ('Fold change',   'Corrected'),
    'AW':    ('Fold change.1', 'Corrected.1'),
    'PA':    ('Fold change.2', 'Corrected.2'),
}
CONDITIONS = ['Intox', 'AW', 'PA']

# ── Colors ────────────────────────────────────────────────────────────────────
FILL_UP   = PatternFill(fill_type='solid', fgColor='FADADD')   # pastel red
FILL_DOWN = PatternFill(fill_type='solid', fgColor='DAE8FA')   # pastel blue
FILL_MISS = PatternFill(fill_type='solid', fgColor='F0F0F0')   # grey
FILL_NS   = PatternFill(fill_type='solid', fgColor='FFFFFF')   # white
FILL_HEAD = PatternFill(fill_type='solid', fgColor='D9D9D9')   # header grey
FILL_FRAC = PatternFill(fill_type='solid', fgColor='E8E8F0')   # fraction row

FONT_HEAD  = Font(bold=True, name='Arial', size=11)
FONT_FRAC  = Font(bold=True, name='Arial', size=10, italic=True)
FONT_GENE  = Font(bold=True, name='Arial', size=10)
FONT_CELL  = Font(name='Arial', size=10)
FONT_NS    = Font(name='Arial', size=10, color='888888')
FONT_MISS  = Font(name='Arial', size=10, color='AAAAAA')

THIN = Side(style='thin', color='CCCCCC')
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=False)
LEFT   = Alignment(horizontal='left',   vertical='center')

def load_sheet(path, sheet):
    df = pd.read_excel(path, sheet_name=sheet)
    df['Gene symbol'] = df['Gene symbol'].astype(str).str.strip()
    if 'Filter' in df.columns:
        df = df[df['Filter'].isin(['Keep', 'Review'])]
    return df

# ── Load all sheets ────────────────────────────────────────────────────────────
sheets = {}
for frac_name, _ in FRACTIONS:
    sheets[frac_name] = load_sheet(DATA_FILE, frac_name)
    print(f'Loaded {frac_name}: {len(sheets[frac_name])} rows')

# ── Collect data ───────────────────────────────────────────────────────────────
# results[gene][fraction][condition] = (fc, corr, direction)
# direction: 'up', 'down', 'ns', or None (not detected)
results = {}
for display, rat_sym in PDC_GENES:
    results[display] = {}
    for frac_name, _ in FRACTIONS:
        df = sheets[frac_name]
        match = df[df['Gene symbol'].str.lower() == rat_sym.lower()]
        results[display][frac_name] = {}
        if match.empty:
            for cond in CONDITIONS:
                results[display][frac_name][cond] = None  # not detected
            print(f'  {display} ({rat_sym}): NOT FOUND in {frac_name}')
        else:
            r = match.iloc[0]
            for cond in CONDITIONS:
                fc_col, corr_col = COND_COLS[cond]
                fc   = pd.to_numeric(r.get(fc_col,   np.nan), errors='coerce')
                corr = pd.to_numeric(r.get(corr_col, np.nan), errors='coerce')
                if pd.isna(fc):
                    direction = None
                elif not pd.isna(corr) and corr > CORR_THRESH and fc > FC_THRESH:
                    direction = 'up'
                elif not pd.isna(corr) and corr > CORR_THRESH and fc < -FC_THRESH:
                    direction = 'down'
                else:
                    direction = 'ns'
                results[display][frac_name][cond] = (fc, corr, direction)
            print(f'  {display} ({rat_sym}): found in {frac_name}')

# ── Build Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'PDC Components'

# Column layout:
# Col 1: Gene | Col 2: Fraction | then for each condition: FC, Corr p
# Conditions: Intox (3,4), AW (5,6), PA (7,8)
GENE_COL = 1
FRAC_COL = 2
COL_START = 3   # first data column

# Header row 1: merged condition headers
ws.merge_cells(start_row=1, start_column=COL_START,   end_row=1, end_column=COL_START+1)
ws.merge_cells(start_row=1, start_column=COL_START+2, end_row=1, end_column=COL_START+3)
ws.merge_cells(start_row=1, start_column=COL_START+4, end_row=1, end_column=COL_START+5)

for i, cond in enumerate(CONDITIONS):
    col = COL_START + i * 2
    c = ws.cell(row=1, column=col, value=cond)
    c.font      = FONT_HEAD
    c.fill      = FILL_HEAD
    c.alignment = CENTER
    c.border    = BORDER_THIN

# Header row 1 for Gene and Fraction
for col, val in [(GENE_COL, 'Gene'), (FRAC_COL, 'Fraction')]:
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    c = ws.cell(row=1, column=col, value=val)
    c.font      = FONT_HEAD
    c.fill      = FILL_HEAD
    c.alignment = CENTER
    c.border    = BORDER_THIN

# Header row 2: FC / Corr p subheaders
for i in range(len(CONDITIONS)):
    col_fc   = COL_START + i * 2
    col_corr = col_fc + 1
    for col, val in [(col_fc, 'FC (log2)'), (col_corr, 'Corr. p (−log2)')]:
        c = ws.cell(row=2, column=col, value=val)
        c.font      = FONT_HEAD
        c.fill      = FILL_HEAD
        c.alignment = CENTER
        c.border    = BORDER_THIN

# Data rows
row = 3
for display, _ in PDC_GENES:
    gene_start_row = row
    for frac_name, _ in FRACTIONS:
        ws.cell(row=row, column=FRAC_COL, value=frac_name).font = FONT_CELL
        ws.cell(row=row, column=FRAC_COL).alignment = LEFT
        ws.cell(row=row, column=FRAC_COL).border = BORDER_THIN

        for i, cond in enumerate(CONDITIONS):
            col_fc   = COL_START + i * 2
            col_corr = col_fc + 1
            val = results[display][frac_name][cond]

            if val is None:
                # Not detected
                for col in (col_fc, col_corr):
                    c = ws.cell(row=row, column=col, value='—')
                    c.fill      = FILL_MISS
                    c.font      = FONT_MISS
                    c.alignment = CENTER
                    c.border    = BORDER_THIN
            else:
                fc, corr, direction = val
                fc_str   = f'{fc:+.2f}' if not pd.isna(fc)   else '—'
                corr_str = f'{corr:.1f}' if not pd.isna(corr) else '—'

                if direction == 'up':
                    fill = FILL_UP
                    font = Font(name='Arial', size=10, bold=True)
                elif direction == 'down':
                    fill = FILL_DOWN
                    font = Font(name='Arial', size=10, bold=True)
                else:
                    fill = FILL_NS
                    font = FONT_NS

                c_fc = ws.cell(row=row, column=col_fc, value=fc_str)
                c_fc.fill = fill; c_fc.font = font
                c_fc.alignment = CENTER; c_fc.border = BORDER_THIN

                c_corr = ws.cell(row=row, column=col_corr, value=corr_str)
                c_corr.fill = fill; c_corr.font = font
                c_corr.alignment = CENTER; c_corr.border = BORDER_THIN

        row += 1

    # Merge gene column across its 4 fraction rows
    if row - 1 > gene_start_row:
        ws.merge_cells(start_row=gene_start_row, start_column=GENE_COL,
                       end_row=row - 1,           end_column=GENE_COL)
    c = ws.cell(row=gene_start_row, column=GENE_COL, value=display)
    c.font      = FONT_GENE
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = BORDER_THIN

    # Add a thin divider row between genes (except after last)
    if display != PDC_GENES[-1][0]:
        for col in range(1, COL_START + len(CONDITIONS) * 2):
            ws.cell(row=row, column=col).fill = PatternFill(fill_type='solid', fgColor='E0E0E0')
        row += 1

# ── Column widths ──────────────────────────────────────────────────────────────
ws.column_dimensions[get_column_letter(GENE_COL)].width = 12
ws.column_dimensions[get_column_letter(FRAC_COL)].width = 18
for i in range(len(CONDITIONS)):
    ws.column_dimensions[get_column_letter(COL_START + i * 2)].width     = 12
    ws.column_dimensions[get_column_letter(COL_START + i * 2 + 1)].width = 16

ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 18

# ── Legend sheet ───────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Legend')
legend = [
    ('Color', 'Meaning'),
    ('Red (bold)',  'Upregulated: FC > +0.5 AND corr. −log2 p > 3.3'),
    ('Blue (bold)', 'Downregulated: FC < −0.5 AND corr. −log2 p > 3.3'),
    ('Grey',        'Not detected in this fraction'),
    ('White',       'Detected but not significant'),
    ('',            ''),
    ('Note:', 'AW-M-3 excluded. Chromatin: Keep+Review only.'),
    ('',      'FC = log2 fold change vs Naïve.'),
    ('',      'Corr. p = −log2 BH-corrected p-value.'),
    ('',      ''),
    ('PDC subunits:',''),
    ('PDHA1', 'Pyruvate dehydrogenase E1 alpha subunit'),
    ('PDHB',  'Pyruvate dehydrogenase E1 beta subunit'),
    ('DLAT',  'Dihydrolipoamide S-acetyltransferase (E2)'),
    ('DLD',   'Dihydrolipoamide dehydrogenase (E3)'),
    ('PDHX',  'Pyruvate dehydrogenase complex component X'),
]
fills_leg = [FILL_HEAD, FILL_UP, FILL_DOWN, FILL_MISS, FILL_NS]
for r, (a, b) in enumerate(legend, 1):
    ws2.cell(row=r, column=1, value=a).font = Font(bold=True, name='Arial', size=10)
    ws2.cell(row=r, column=2, value=b).font = Font(name='Arial', size=10)
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 55

wb.save(OUT_XLSX)
print(f'Saved: {OUT_XLSX}')
