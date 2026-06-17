"""
Export the proteins used in both heatmap panel figures to Excel.
One sheet per compartment × version combination.
Includes: Gene symbol, Protein name, and FC + significance for each condition.
AW-M-3 excluded from Acute Withdrawal. Chromatin: Keep+Review only.

Output: Heatmap_Proteins.xlsx
"""

import pandas as pd
import numpy as np
from scipy import stats
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE = '../EDIT  Philipp Alcohol proteome Simplified - Jan 2024  copy.xlsx'
OUT  = 'Heatmap_Proteins.xlsx'

CORR_THRESH = 3.3
FC_THRESH   = 0.5

CONDITIONS = [
    ('Intoxication',          ['I-F-1','I-F-2','I-F-3','I-M-1','I-M-2']),
    ('Acute Withdrawal',      ['AW-F-1','AW-F-2','AW-M-1','AW-M-2']),
    ('Protracted Abstinence', ['PA-F-1','PA-F-2','PA-F-3','PA-M-1','PA-M-2']),
]
NAIVE_SUFFIXES = ['N-F-1','N-F-2','N-F-3','N-M-1','N-M-2']

SHEETS = [
    ('Membrane',        'Membrane',        'Memb_',  'all'),
    ('Cytosol',         'Cytosol',         'Cyto_',  'all'),
    ('Chromatin',       'Chromatin',       'Chrom_', 'keep_review'),
    ('Soluble Nuclear', 'Soluble nuclear', 'Nuc_',   'all'),
]

def calc_stats(df, naive_cols, cond_cols):
    naive = df[naive_cols].apply(pd.to_numeric, errors='coerce')
    cond  = df[cond_cols].apply(pd.to_numeric, errors='coerce')
    fc    = cond.mean(axis=1) - naive.mean(axis=1)
    pvs   = []
    for i in range(len(df)):
        n = naive.iloc[i].dropna().values
        c = cond.iloc[i].dropna().values
        pvs.append(stats.ttest_ind(n, c, equal_var=False)[1]
                   if len(n) >= 2 and len(c) >= 2 else np.nan)
    pvs   = pd.Series(pvs, index=df.index)
    valid = pvs.notna()
    ranks = pvs[valid].rank(ascending=False)
    corr  = pd.Series(np.nan, index=df.index)
    corr[valid] = -np.log2(np.minimum(1, pvs[valid] * valid.sum() / ranks))
    return fc, corr

# ── Collect data ───────────────────────────────────────────────────────────────
writer = pd.ExcelWriter(OUT, engine='openpyxl')

summary_rows = []   # for the Summary sheet

for (disp, sheet, prefix, fmode) in SHEETS:
    raw = pd.read_excel(FILE, sheet_name=sheet)
    if fmode == 'keep_review' and 'Filter' in raw.columns:
        df = raw[raw['Filter'].isin(['Keep','Review'])].reset_index(drop=True)
    else:
        df = raw.copy().reset_index(drop=True)

    naive_cols = [c for c in df.columns if any(prefix+s in str(c) for s in NAIVE_SUFFIXES)]

    fc_map   = {}
    corr_map = {}
    for cname, csuf in CONDITIONS:
        cc = [c for c in df.columns if any(prefix+s in str(c) for s in csuf)]
        if cc:
            fc, corr = calc_stats(df, naive_cols, cc)
            fc_map[cname]   = fc
            corr_map[cname] = corr

    # Build significance masks
    sig_per_cond = {}
    for cname, _ in CONDITIONS:
        if cname in fc_map:
            sig_per_cond[cname] = (
                corr_map[cname].notna() & fc_map[cname].notna() &
                (corr_map[cname] > CORR_THRESH) & (fc_map[cname].abs() > FC_THRESH)
            )

    union_mask  = pd.Series(False, index=df.index)
    allsig_mask = pd.Series(True,  index=df.index)
    for cname, _ in CONDITIONS:
        if cname in sig_per_cond:
            union_mask  = union_mask  | sig_per_cond[cname]
            allsig_mask = allsig_mask & sig_per_cond[cname]

    # ── Build output dataframe ──────────────────────────────────────────────
    # Try to find gene symbol and protein name columns
    gene_col  = next((c for c in df.columns if 'gene' in str(c).lower()), None)
    prot_col  = next((c for c in df.columns
                      if any(x in str(c).lower() for x in ['protein name','protein description','name'])
                      and 'gene' not in str(c).lower()), None)

    for mask, label in [(union_mask, 'Union'), (allsig_mask, 'AllSig')]:
        df_sub = df[mask].reset_index(drop=True)
        n = len(df_sub)

        rows = {}
        if gene_col:
            rows['Gene Symbol'] = df_sub[gene_col].astype(str).values
        if prot_col:
            rows['Protein Name'] = df_sub[prot_col].astype(str).values

        naive_v = df_sub[naive_cols].apply(pd.to_numeric, errors='coerce')

        for cname, csuf in CONDITIONS:
            cc = [c for c in df.columns if any(prefix+s in str(c) for s in csuf)]
            if cc:
                cond_v = df_sub[cc].apply(pd.to_numeric, errors='coerce')
                fc_vals   = (cond_v.mean(axis=1) - naive_v.mean(axis=1)).values
                corr_vals = corr_map[cname][mask].reset_index(drop=True).values
                sig_vals  = sig_per_cond[cname][mask].reset_index(drop=True).values

                short = cname.replace('Protracted Abstinence','PA')\
                             .replace('Acute Withdrawal','AW')\
                             .replace('Intoxication','Intox')
                rows[f'FC_{short}']      = np.round(fc_vals, 4)
                rows[f'CorrP_{short}']   = np.round(corr_vals, 4)
                rows[f'Sig_{short}']     = ['YES' if s else 'no' for s in sig_vals]

        out_df = pd.DataFrame(rows)

        # Sort by AW FC if available
        if 'FC_AW' in out_df.columns:
            out_df = out_df.sort_values('FC_AW').reset_index(drop=True)

        sheet_name = f'{disp[:12]}_{label}'   # Excel sheet names max 31 chars
        out_df.to_excel(writer, sheet_name=sheet_name, index=False)

        summary_rows.append({
            'Compartment':  disp,
            'Version':      label,
            'N Proteins':   n,
            'Filter':       'Keep+Review only' if fmode == 'keep_review' else 'All rows',
            'Sheet':        sheet_name,
        })
        print(f'  {disp} {label}: {n} proteins → sheet "{sheet_name}"')

# ── Summary sheet ──────────────────────────────────────────────────────────────
summary_df = pd.DataFrame(summary_rows)
summary_df.to_excel(writer, sheet_name='Summary', index=False)

writer.close()
import time; time.sleep(1)   # ensure file is fully flushed before re-opening

# ── Light formatting ───────────────────────────────────────────────────────────
from openpyxl import load_workbook
wb = load_workbook(OUT)

HEADER_FILL = PatternFill('solid', fgColor='2166AC')   # dark blue header
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
ALT_FILL    = PatternFill('solid', fgColor='EEF4FB')   # light blue alternate rows
BORDER      = Border(bottom=Side(style='thin', color='CCCCCC'))

for ws in wb.worksheets:
    # Header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows: alternate shading + auto-width
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
            # Highlight significant YES cells in red text
            if str(cell.value) == 'YES':
                cell.font = Font(bold=True, color='B22222', name='Arial', size=10)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 30)

wb.save(OUT)
print(f'\nSaved: {OUT}')
